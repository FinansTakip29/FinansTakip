from io import BytesIO
import calendar
import json
from decimal import Decimal, InvalidOperation, ROUND_CEILING
from datetime import date, timedelta
from pathlib import Path
from xml.sax.saxutils import escape

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.templatetags.static import static
from django.utils import timezone
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import (
    FINANS_KISISEL,
    FINANS_TURU_SECENEKLERI,
    BirikimHedefi,
    ButceHedefi,
    Gelir,
    Gider,
    Kategori,
    KategoriButcesi,
    OdemeDonemi,
    TekrarlayanOdeme,
)


def health(request):
    return JsonResponse({"status": "ok"})


def favicon(request):
    svg = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64"><rect width="64" height="64" rx="12" fill="#2563eb"/><path fill="#fff" d="M18 20h28a6 6 0 0 1 6 6v20a6 6 0 0 1-6 6H18a6 6 0 0 1-6-6V26a6 6 0 0 1 6-6Zm0 8v18h28V28H18Zm23 7h6v6h-6z"/><path fill="#93c5fd" d="M20 12h22a4 4 0 0 1 4 4v4H18v-6a2 2 0 0 1 2-2Z"/></svg>"""
    response = HttpResponse(svg, content_type="image/svg+xml")
    response["Cache-Control"] = "public, max-age=86400"
    return response


def manifest(request):
    icon_sizes = [72, 96, 128, 144, 152, 192, 384, 512]
    data = {
        "name": settings.PWA_APP_NAME,
        "short_name": settings.PWA_APP_SHORT_NAME,
        "description": settings.PWA_APP_DESCRIPTION,
        "start_url": "/",
        "scope": "/",
        "display": settings.PWA_APP_DISPLAY,
        "orientation": settings.PWA_APP_ORIENTATION,
        "theme_color": settings.PWA_APP_THEME_COLOR,
        "background_color": settings.PWA_APP_BACKGROUND_COLOR,
        "icons": [
            {
                "src": static(f"icons/icon-{size}x{size}.png"),
                "sizes": f"{size}x{size}",
                "type": "image/png",
                "purpose": "any maskable",
            }
            for size in icon_sizes
        ],
        "shortcuts": [
            {
                "name": "Hızlı Harcama Ekle",
                "short_name": "Harcama",
                "description": "Hızlı gider ekle",
                "url": "/?quick_add=1",
                "icons": [
                    {
                        "src": static("icons/icon-192x192.png"),
                        "sizes": "192x192",
                        "type": "image/png",
                    }
                ],
            }
        ],
    }
    return JsonResponse(data, content_type="application/manifest+json")


def service_worker(request):
    cache_name = "finanstakip-pwa-v4"
    static_assets = [
        "/offline/",
        "/favicon.ico",
        "/manifest.json",
        static("css/app.css"),
        static("js/app.js"),
        *[static(f"icons/icon-{size}x{size}.png") for size in [72, 96, 128, 144, 152, 192, 384, 512]],
    ]
    script = f"""
const CACHE_NAME = {json.dumps(cache_name)};
const OFFLINE_URL = "/offline/";
const STATIC_ASSETS = {json.dumps(static_assets)};
const AUTH_PATHS = ["/giris/", "/kayit/", "/cikis/", "/admin/"];
const NEVER_CACHE_PATHS = [
    ...AUTH_PATHS,
    "/service-worker.js",
    "/gelir-ekle/",
    "/gider-ekle/",
    "/hizli-gider-ekle/",
    "/butce-hedefi/",
    "/kategori-butceleri/",
    "/kategori-butcesi-",
    "/birikim-hedefleri/",
    "/birikim-hedefi-",
    "/kategoriler/",
    "/tekrarlayan-odemeler/",
    "/yedekleme/",
    "/raporlar/",
];
const PRECACHE_URLS = new Set(STATIC_ASSETS);

function isSameOrigin(url) {{
    return url.origin === self.location.origin;
}}

function startsWithAny(url, paths) {{
    return paths.some((path) => url.pathname.startsWith(path));
}}

function shouldBypassCache(request, url) {{
    if (request.method !== "GET") {{
        return true;
    }}
    if (!isSameOrigin(url)) {{
        return true;
    }}
    return startsWithAny(url, NEVER_CACHE_PATHS);
}}

function isPrecacheAsset(url) {{
    return PRECACHE_URLS.has(url.pathname);
}}

self.addEventListener("install", (event) => {{
    event.waitUntil(
        caches.open(CACHE_NAME)
            .then((cache) => cache.addAll(STATIC_ASSETS))
            .then(() => self.skipWaiting())
    );
}});

self.addEventListener("activate", (event) => {{
    event.waitUntil(
        caches.keys()
            .then((keys) => Promise.all(keys.filter((key) => key !== CACHE_NAME).map((key) => caches.delete(key))))
            .then(() => self.clients.claim())
    );
}});

self.addEventListener("fetch", (event) => {{
    const request = event.request;
    const url = new URL(request.url);
    if (shouldBypassCache(request, url)) {{
        return;
    }}

    if (request.mode === "navigate") {{
        event.respondWith(
            fetch(request)
                .catch(() => caches.match(OFFLINE_URL))
        );
        return;
    }}

    if (!isPrecacheAsset(url)) {{
        event.respondWith(fetch(request));
        return;
    }}

    event.respondWith(
        caches.match(request)
            .then((cached) => cached || fetch(request).then((response) => {{
                if (response && response.status === 200) {{
                    const copy = response.clone();
                    caches.open(CACHE_NAME).then((cache) => cache.put(request, copy));
                }}
                return response;
            }}))
            .catch(() => caches.match(OFFLINE_URL))
    );
}});

self.addEventListener("push", (event) => {{
    const data = event.data ? event.data.json() : {{}};
    const title = data.title || "FinansTakip";
    const options = {{
        body: data.body || "Yeni finans bildirimi var.",
        icon: "/static/icons/icon-192x192.png",
        badge: "/static/icons/icon-96x96.png",
        data: {{ url: data.url || "/" }}
    }};
    event.waitUntil(self.registration.showNotification(title, options));
}});

self.addEventListener("notificationclick", (event) => {{
    event.notification.close();
    const targetUrl = event.notification.data && event.notification.data.url ? event.notification.data.url : "/";
    event.waitUntil(clients.openWindow(targetUrl));
}});
"""
    response = HttpResponse(script, content_type="application/javascript")
    response["Service-Worker-Allowed"] = "/"
    response["Cache-Control"] = "no-cache"
    return response


def offline(request):
    return render(request, "offline.html")


def _secilen_finans_turu(veri):
    finans_turu = veri.get("finans_turu") or FINANS_KISISEL
    gecerli_turler = [deger for deger, _ in FINANS_TURU_SECENEKLERI]
    if finans_turu not in gecerli_turler:
        return FINANS_KISISEL
    return finans_turu


def _finans_turu_context(finans_turu):
    return {
        "finans_turleri": FINANS_TURU_SECENEKLERI,
        "secilen_finans_turu": finans_turu,
    }


def _ay_sonu(yil, ay):
    return calendar.monthrange(yil, ay)[1]


def _ay_ekle(tarih, ay_sayisi):
    toplam_ay = tarih.month - 1 + ay_sayisi
    yil = tarih.year + toplam_ay // 12
    ay = toplam_ay % 12 + 1
    gun = min(tarih.day, _ay_sonu(yil, ay))
    return tarih.replace(year=yil, month=ay, day=gun)


def _yil_ekle(tarih, yil_sayisi):
    yil = tarih.year + yil_sayisi
    gun = min(tarih.day, _ay_sonu(yil, tarih.month))
    return tarih.replace(year=yil, day=gun)


def _sonraki_odeme_tarihi(odeme, referans_tarihi):
    tarih = odeme.baslangic_tarihi
    aralik = max(odeme.tekrar_araligi or 1, 1)

    if tarih > referans_tarihi:
        return tarih

    while tarih < referans_tarihi:
        if odeme.tekrar_turu == TekrarlayanOdeme.GUNLUK:
            tarih = tarih + timedelta(days=aralik)
        elif odeme.tekrar_turu == TekrarlayanOdeme.HAFTALIK:
            tarih = tarih + timedelta(weeks=aralik)
        elif odeme.tekrar_turu == TekrarlayanOdeme.OZEL:
            tarih = tarih + timedelta(days=aralik)
        elif odeme.tekrar_turu == TekrarlayanOdeme.YILLIK:
            tarih = _yil_ekle(tarih, aralik)
        else:
            tarih = _ay_ekle(tarih, aralik)

    return tarih


def _bekleyen_odeme_tarihi(odeme):
    if odeme.son_olusturma_tarihi:
        referans_tarihi = odeme.son_olusturma_tarihi + timedelta(days=1)
    else:
        referans_tarihi = odeme.baslangic_tarihi

    return _sonraki_odeme_tarihi(odeme, referans_tarihi)


def _tekrarlayan_odeme_gideri_olustur(odeme, vade_tarihi):
    if odeme.son_gider and odeme.son_olusturma_tarihi == vade_tarihi:
        gider = odeme.son_gider
    else:
        gider = Gider.objects.create(
            kullanici=odeme.kullanici,
            finans_turu=odeme.finans_turu,
            tarih=vade_tarihi,
            aciklama=odeme.aciklama or odeme.odeme_adi,
            tutar=odeme.tutar,
            kategori=odeme.kategori.ad,
        )

    odeme.son_olusturma_tarihi = vade_tarihi
    odeme.son_gider = gider
    odeme.odeme_durumu = TekrarlayanOdeme.ODENDI
    odeme.save(update_fields=["son_olusturma_tarihi", "son_gider", "odeme_durumu"])
    return gider


def _odeme_donemi_gideri_olustur(donem):
    odeme = donem.tekrarlayan_odeme
    gider = Gider.objects.create(
        kullanici=odeme.kullanici,
        finans_turu=odeme.finans_turu,
        tarih=donem.vade_tarihi,
        aciklama=odeme.aciklama or odeme.odeme_adi,
        tutar=odeme.tutar,
        kategori=odeme.kategori.ad,
    )
    donem.durum = OdemeDonemi.ODENDI
    donem.save(update_fields=["durum"])
    odeme.son_olusturma_tarihi = donem.vade_tarihi
    odeme.son_gider = gider
    odeme.odeme_durumu = TekrarlayanOdeme.ODENDI
    odeme.save(update_fields=["son_olusturma_tarihi", "son_gider", "odeme_durumu"])
    return gider


def _donem_durum_verisi(donem, bugun):
    if donem.durum == OdemeDonemi.ODENDI:
        return {"deger": OdemeDonemi.ODENDI, "etiket": "Ödendi", "renk": "success"}
    if donem.durum == OdemeDonemi.IPTAL:
        return {"deger": OdemeDonemi.IPTAL, "etiket": "İptal Edildi", "renk": "secondary"}
    if donem.vade_tarihi < bugun:
        return {"deger": OdemeDonemi.GECIKTI, "etiket": "Gecikti", "renk": "danger"}
    return {"deger": OdemeDonemi.BEKLIYOR, "etiket": "Bekliyor", "renk": "warning"}


def _donem_durumu_guncelle(donem, bugun):
    durum = _donem_durum_verisi(donem, bugun)
    if donem.durum in [OdemeDonemi.ODENDI, OdemeDonemi.IPTAL]:
        return durum
    if donem.durum != durum["deger"]:
        donem.durum = durum["deger"]
        donem.save(update_fields=["durum"])
    return durum


def _donem_vade_tarihi(odeme, yil, ay):
    ay_baslangici = date(yil, ay, 1)
    ay_bitisi = ay_baslangici.replace(day=_ay_sonu(yil, ay))
    if odeme.baslangic_tarihi > ay_bitisi:
        return None

    referans_tarihi = max(ay_baslangici, odeme.baslangic_tarihi)
    vade_tarihi = _sonraki_odeme_tarihi(odeme, referans_tarihi)
    if vade_tarihi <= ay_bitisi:
        return vade_tarihi
    return None


def _odeme_donemlerini_olustur(kullanici, finans_turu=None):
    bugun = timezone.now().date()
    ay_bitisi = bugun.replace(day=_ay_sonu(bugun.year, bugun.month))
    odemeler = TekrarlayanOdeme.objects.filter(
        kullanici=kullanici,
        aktif=True,
        baslangic_tarihi__lte=ay_bitisi,
    ).select_related("kategori")

    if finans_turu:
        odemeler = odemeler.filter(finans_turu=finans_turu)

    for odeme in odemeler:
        yil = odeme.baslangic_tarihi.year
        ay = odeme.baslangic_tarihi.month

        while (yil, ay) <= (bugun.year, bugun.month):
            vade_tarihi = _donem_vade_tarihi(odeme, yil, ay)
            if vade_tarihi:
                OdemeDonemi.objects.get_or_create(
                    tekrarlayan_odeme=odeme,
                    donem_yil=yil,
                    donem_ay=ay,
                    defaults={
                        "vade_tarihi": vade_tarihi,
                        "durum": OdemeDonemi.BEKLIYOR,
                    },
                )

            ay += 1
            if ay > 12:
                ay = 1
                yil += 1


def _odeme_durumu_verisi(odeme, vade_tarihi, bugun):
    if not odeme.aktif or odeme.odeme_durumu == TekrarlayanOdeme.IPTAL:
        return {
            "deger": TekrarlayanOdeme.IPTAL,
            "etiket": "İptal edildi",
            "renk": "secondary",
        }

    if odeme.son_olusturma_tarihi == vade_tarihi and odeme.son_gider_id:
        return {
            "deger": TekrarlayanOdeme.ODENDI,
            "etiket": "Ödendi",
            "renk": "success",
        }

    if vade_tarihi < bugun:
        return {
            "deger": TekrarlayanOdeme.GECIKTI,
            "etiket": "Gecikti",
            "renk": "danger",
        }

    return {
        "deger": TekrarlayanOdeme.BEKLIYOR,
        "etiket": "Bekliyor",
        "renk": "warning",
    }


def _tekrarlayan_odemeleri_olustur(kullanici):
    _odeme_donemlerini_olustur(kullanici)


def _tekrarlayan_odeme_verileri(kullanici, finans_turu):
    bugun = timezone.now().date()
    ay_baslangici = bugun.replace(day=1)
    ay_bitisi = bugun.replace(day=_ay_sonu(bugun.year, bugun.month))
    _odeme_donemlerini_olustur(kullanici, finans_turu)

    donemler = OdemeDonemi.objects.filter(
        tekrarlayan_odeme__kullanici=kullanici,
        tekrarlayan_odeme__finans_turu=finans_turu,
        vade_tarihi__gte=ay_baslangici,
        vade_tarihi__lte=ay_bitisi,
    ).select_related("tekrarlayan_odeme", "tekrarlayan_odeme__kategori").order_by("vade_tarihi")

    bu_ay_odemeler = []
    yaklasan_odemeler = []
    geciken_odemeler = []

    for donem in donemler:
        durum = _donem_durumu_guncelle(donem, bugun)
        kayit = {
            "donem": donem,
            "odeme": donem.tekrarlayan_odeme,
            "vade_tarihi": donem.vade_tarihi,
            "durum": durum,
        }
        bu_ay_odemeler.append(kayit)

        if durum["deger"] == OdemeDonemi.GECIKTI:
            geciken_odemeler.append(kayit)
        elif durum["deger"] == OdemeDonemi.BEKLIYOR:
            yaklasan_odemeler.append(kayit)

    return {
        "bu_ay_tekrarlayan_odemeler": bu_ay_odemeler,
        "yaklasan_odemeler": yaklasan_odemeler,
        "geciken_odemeler": geciken_odemeler,
    }


def _aylik_butce_verileri(kullanici, finans_turu):
    bugun = timezone.now().date()
    yil = bugun.year
    ay = bugun.month

    bu_ay_gider = Gider.objects.filter(
        kullanici=kullanici,
        finans_turu=finans_turu,
        tarih__year=yil,
        tarih__month=ay,
    ).aggregate(Sum("tutar"))["tutar__sum"] or 0

    butce_hedefi = ButceHedefi.objects.filter(
        kullanici=kullanici,
        finans_turu=finans_turu,
        yil=yil,
        ay=ay,
    ).first()

    hedef_tutar = butce_hedefi.hedef_tutar if butce_hedefi else None
    kalan_butce = hedef_tutar - bu_ay_gider if hedef_tutar is not None else None
    butce_uyari = None

    if hedef_tutar and hedef_tutar > 0:
        kullanim_orani = (bu_ay_gider / hedef_tutar) * 100
        if bu_ay_gider > hedef_tutar:
            butce_uyari = "danger"
        elif kullanim_orani > 80:
            butce_uyari = "warning"
    else:
        kullanim_orani = 0

    return {
        "butce_hedefi": butce_hedefi,
        "butce_yil": yil,
        "butce_ay": ay,
        "aylik_butce_hedefi": hedef_tutar,
        "bu_ay_gider": bu_ay_gider,
        "kalan_butce": kalan_butce,
        "butce_uyari": butce_uyari,
        "butce_kullanim_orani": round(float(kullanim_orani), 2),
        **_finans_turu_context(finans_turu),
    }


def _kategori_butce_verileri(kullanici, finans_turu):
    bugun = timezone.now().date()
    butceler = KategoriButcesi.objects.filter(
        kullanici=kullanici,
        finans_turu=finans_turu,
        yil=bugun.year,
        ay=bugun.month,
    ).select_related("kategori").order_by("kategori__ad")

    kayitlar = []
    asilan_sayisi = 0
    uyarilar = []

    for butce in butceler:
        harcanan = Gider.objects.filter(
            kullanici=kullanici,
            finans_turu=finans_turu,
            kategori=butce.kategori.ad,
            tarih__year=butce.yil,
            tarih__month=butce.ay,
        ).aggregate(Sum("tutar"))["tutar__sum"] or Decimal("0")
        kalan = butce.hedef_tutar - harcanan
        oran = (harcanan / butce.hedef_tutar) * 100 if butce.hedef_tutar else Decimal("0")
        oran_float = round(float(oran), 2)

        if harcanan > butce.hedef_tutar:
            renk = "danger"
            asilan_sayisi += 1
            uyarilar.append(f"{butce.kategori.ad} bütçesi aşıldı.")
        elif oran_float >= 80:
            renk = "warning"
            uyarilar.append(f"{butce.kategori.ad} bütçesinin %80'i aşıldı.")
        else:
            renk = "success"

        kayitlar.append({
            "butce": butce,
            "harcanan": harcanan,
            "kalan": kalan,
            "oran": oran_float,
            "progress_orani": min(oran_float, 100),
            "renk": renk,
        })

    return {
        "kategori_butce_kayitlari": kayitlar,
        "kategori_butce_uyarilari": uyarilar,
        "kategori_butce_asim_sayisi": asilan_sayisi,
    }


def _birikim_hedefi_verileri(kullanici, finans_turu):
    hedefler = BirikimHedefi.objects.filter(
        kullanici=kullanici,
        finans_turu=finans_turu,
        aktif=True,
    ).order_by("hedef_tarihi", "hedef_adi")

    kayitlar = []
    for hedef in hedefler:
        kalan = max(hedef.hedef_tutar - hedef.mevcut_tutar, Decimal("0"))
        oran = (hedef.mevcut_tutar / hedef.hedef_tutar) * 100 if hedef.hedef_tutar else Decimal("0")
        oran_float = round(float(oran), 2)
        tahmini_ay = None
        if hedef.aylik_katki and hedef.aylik_katki > 0 and kalan > 0:
            tahmini_ay = int((kalan / hedef.aylik_katki).to_integral_value(rounding=ROUND_CEILING))

        kayitlar.append({
            "hedef": hedef,
            "kalan": kalan,
            "oran": oran_float,
            "progress_orani": min(oran_float, 100),
            "tahmini_ay": tahmini_ay,
        })

    return {"birikim_hedefi_kayitlari": kayitlar}


def _finansal_tavsiyeler(grafik_verileri, butce_verileri, odeme_verileri, kategori_butce_verileri):
    skor = grafik_verileri["finansal_saglik_puani"]
    tavsiyeler = []

    if skor < 50:
        durum = {"etiket": "Kritik", "renk": "danger"}
        tavsiyeler.append("Tasarruf oranınız kritik seviyede. Bu ay zorunlu olmayan harcamaları azaltmayı deneyebilirsiniz.")
    elif skor < 70:
        durum = {"etiket": "Dikkat", "renk": "warning"}
        tavsiyeler.append("Gelir/gider dengenizi yakından izleyin; bütçe sınırlarına yaklaşan kategorileri kontrol edin.")
    elif skor < 85:
        durum = {"etiket": "İyi", "renk": "success"}
        tavsiyeler.append("Bu ay finansal durumunuz iyi görünüyor. Birikim hedeflerinize düzenli katkı yapabilirsiniz.")
    else:
        durum = {"etiket": "Çok iyi", "renk": "primary"}
        tavsiyeler.append("Finansal sağlığınız çok iyi. Birikim katkınızı artırmayı değerlendirebilirsiniz.")

    if butce_verileri["butce_uyari"] == "warning":
        tavsiyeler.append("Bütçe hedefinizin %80'ini aştınız. Harcamalarınızı kontrol etmeniz iyi olur.")
    elif butce_verileri["butce_uyari"] == "danger":
        tavsiyeler.append("Aylık bütçe hedefiniz aşıldı. Yeni harcamalarda daha seçici davranın.")

    if odeme_verileri["geciken_odemeler"]:
        tavsiyeler.append("Geciken ödemeleriniz var. Öncelikle bu ödemeleri kapatmak finansal skorunuzu iyileştirir.")

    for uyari in kategori_butce_verileri["kategori_butce_uyarilari"][:3]:
        tavsiyeler.append(uyari)

    return {
        "finansal_saglik_durumu": durum,
        "finansal_tavsiyeler": tavsiyeler[:5],
    }


def _rapor_verileri(kullanici, finans_turu, ay=None):
    tum_gelirler = Gelir.objects.filter(kullanici=kullanici, finans_turu=finans_turu)
    tum_giderler = Gider.objects.filter(kullanici=kullanici, finans_turu=finans_turu)

    toplam_gelir = tum_gelirler.aggregate(Sum("tutar"))["tutar__sum"] or 0
    toplam_gider = tum_giderler.aggregate(Sum("tutar"))["tutar__sum"] or 0
    bakiye = toplam_gelir - toplam_gider

    gelirler = tum_gelirler
    giderler = tum_giderler

    import re
    if ay and re.match(r'^\d{4}-\d{2}$', ay):
        yil, ay_numarasi = ay.split("-")
        gelirler = gelirler.filter(tarih__year=yil, tarih__month=ay_numarasi)
        giderler = giderler.filter(tarih__year=yil, tarih__month=ay_numarasi)
    else:
        ay = None

    aylik_gelir = gelirler.aggregate(Sum("tutar"))["tutar__sum"] or 0
    aylik_gider = giderler.aggregate(Sum("tutar"))["tutar__sum"] or 0
    aylik_bakiye = aylik_gelir - aylik_gider

    kategori_giderler = giderler.values("kategori").annotate(toplam=Sum("tutar"))

    kategori_adlari = []
    kategori_tutarlari = []
    kategori_detaylari = []

    for item in kategori_giderler:
        kategori_adlari.append(item["kategori"])
        kategori_tutarlari.append(float(item["toplam"]))
        kategori_detaylari.append({
            "ad": item["kategori"],
            "tutar": item["toplam"]
        })

    return {
        "toplam_gelir": toplam_gelir,
        "toplam_gider": toplam_gider,
        "bakiye": bakiye,
        "kategori_adlari": kategori_adlari,
        "kategori_tutarlari": kategori_tutarlari,
        "kategori_detaylari": kategori_detaylari,
        "aylik_gelir": aylik_gelir,
        "aylik_gider": aylik_gider,
        "aylik_bakiye": aylik_bakiye,
        "secilen_ay": ay,
        **_finans_turu_context(finans_turu),
    }


def _dashboard_grafik_verileri(kullanici, finans_turu, butce_verileri, odeme_verileri, kategori_butce_verileri=None):
    bugun = timezone.now().date()
    ay_baslangici = bugun.replace(day=1)
    onceki_ay_baslangici = _ay_ekle(ay_baslangici, -1)

    aylar = []
    for indeks in range(11, -1, -1):
        ay_tarihi = _ay_ekle(ay_baslangici, -indeks)
        aylar.append((ay_tarihi.year, ay_tarihi.month, f"{ay_tarihi.month:02d}/{ay_tarihi.year}"))

    trend_gelirleri = []
    trend_giderleri = []

    for yil, ay, _ in aylar:
        aylik_gelir = Gelir.objects.filter(
            kullanici=kullanici,
            finans_turu=finans_turu,
            tarih__year=yil,
            tarih__month=ay,
        ).aggregate(Sum("tutar"))["tutar__sum"] or 0
        aylik_gider = Gider.objects.filter(
            kullanici=kullanici,
            finans_turu=finans_turu,
            tarih__year=yil,
            tarih__month=ay,
        ).aggregate(Sum("tutar"))["tutar__sum"] or 0
        trend_gelirleri.append(float(aylik_gelir))
        trend_giderleri.append(float(aylik_gider))

    bu_ay_gelir = Gelir.objects.filter(
        kullanici=kullanici,
        finans_turu=finans_turu,
        tarih__year=bugun.year,
        tarih__month=bugun.month,
    ).aggregate(Sum("tutar"))["tutar__sum"] or 0
    bu_ay_gider = butce_verileri["bu_ay_gider"]
    onceki_ay_gelir = Gelir.objects.filter(
        kullanici=kullanici,
        finans_turu=finans_turu,
        tarih__year=onceki_ay_baslangici.year,
        tarih__month=onceki_ay_baslangici.month,
    ).aggregate(Sum("tutar"))["tutar__sum"] or 0
    onceki_ay_gider = Gider.objects.filter(
        kullanici=kullanici,
        finans_turu=finans_turu,
        tarih__year=onceki_ay_baslangici.year,
        tarih__month=onceki_ay_baslangici.month,
    ).aggregate(Sum("tutar"))["tutar__sum"] or 0

    aylik_nakit_akisi = bu_ay_gelir - bu_ay_gider
    onceki_ay_nakit_akisi = onceki_ay_gelir - onceki_ay_gider
    if onceki_ay_nakit_akisi:
        gecen_aya_gore_degisim = ((aylik_nakit_akisi - onceki_ay_nakit_akisi) / abs(onceki_ay_nakit_akisi)) * 100
    else:
        gecen_aya_gore_degisim = 100 if aylik_nakit_akisi > 0 else 0

    tasarruf_orani = (aylik_nakit_akisi / bu_ay_gelir) * 100 if bu_ay_gelir else 0
    butce_kullanim_orani = Decimal(str(butce_verileri["butce_kullanim_orani"]))
    geciken_odeme_sayisi = len(odeme_verileri["geciken_odemeler"])
    kategori_asim_sayisi = 0
    if kategori_butce_verileri:
        kategori_asim_sayisi = kategori_butce_verileri["kategori_butce_asim_sayisi"]
    saglik_puani = 50
    saglik_puani += min(max(float(tasarruf_orani), -30), 30)
    if butce_verileri["aylik_butce_hedefi"]:
        saglik_puani += max(0, 20 - (float(butce_kullanim_orani) / 5))
    saglik_puani -= min(geciken_odeme_sayisi * 10, 30)
    saglik_puani -= min(kategori_asim_sayisi * 8, 24)
    saglik_puani = max(0, min(100, round(saglik_puani)))

    kategori_giderleri = Gider.objects.filter(
        kullanici=kullanici,
        finans_turu=finans_turu,
        tarih__year=bugun.year,
        tarih__month=bugun.month,
    ).values("kategori").annotate(toplam=Sum("tutar")).order_by("-toplam")

    return {
        "dashboard_aylik_gelir": bu_ay_gelir,
        "dashboard_aylik_gider": bu_ay_gider,
        "tasarruf_orani": round(float(tasarruf_orani), 2),
        "finansal_saglik_puani": saglik_puani,
        "gecen_aya_gore_degisim": round(float(gecen_aya_gore_degisim), 2),
        "aylik_nakit_akisi": aylik_nakit_akisi,
        "trend_aylari": [etiket for _, _, etiket in aylar],
        "trend_gelirleri": trend_gelirleri,
        "trend_giderleri": trend_giderleri,
        "kategori_grafik_adlari": [item["kategori"] for item in kategori_giderleri],
        "kategori_grafik_tutarlari": [float(item["toplam"]) for item in kategori_giderleri],
    }


def _date_to_text(value):
    return value.isoformat() if value else None


def _decimal_to_text(value):
    return str(value) if value is not None else None


def _parse_decimal(value):
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _parse_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _backup_payload(kullanici):
    kategoriler = Kategori.objects.filter(kullanici=kullanici).order_by("finans_turu", "tur", "ad")
    gelirler = Gelir.objects.filter(kullanici=kullanici).order_by("tarih", "id")
    giderler = Gider.objects.filter(kullanici=kullanici).order_by("tarih", "id")
    butceler = ButceHedefi.objects.filter(kullanici=kullanici).order_by("yil", "ay", "finans_turu")
    kategori_butceleri = KategoriButcesi.objects.filter(kullanici=kullanici).select_related("kategori").order_by("yil", "ay", "finans_turu", "kategori__ad")
    birikim_hedefleri = BirikimHedefi.objects.filter(kullanici=kullanici).order_by("finans_turu", "hedef_adi")
    odemeler = TekrarlayanOdeme.objects.filter(kullanici=kullanici).select_related("kategori", "son_gider").order_by("id")
    donemler = OdemeDonemi.objects.filter(
        tekrarlayan_odeme__kullanici=kullanici,
    ).select_related("tekrarlayan_odeme").order_by("donem_yil", "donem_ay", "id")

    return {
        "meta": {
            "app": "FinansTakip",
            "version": 1,
            "created_at": timezone.now().isoformat(),
            "username": kullanici.username,
        },
        "kategoriler": [
            {
                "id": kategori.id,
                "finans_turu": kategori.finans_turu,
                "ad": kategori.ad,
                "tur": kategori.tur,
            }
            for kategori in kategoriler
        ],
        "gelirler": [
            {
                "id": gelir.id,
                "finans_turu": gelir.finans_turu,
                "tarih": _date_to_text(gelir.tarih),
                "aciklama": gelir.aciklama,
                "tutar": _decimal_to_text(gelir.tutar),
                "kategori": gelir.kategori,
            }
            for gelir in gelirler
        ],
        "giderler": [
            {
                "id": gider.id,
                "finans_turu": gider.finans_turu,
                "tarih": _date_to_text(gider.tarih),
                "aciklama": gider.aciklama,
                "tutar": _decimal_to_text(gider.tutar),
                "kategori": gider.kategori,
            }
            for gider in giderler
        ],
        "butce_hedefleri": [
            {
                "id": butce.id,
                "finans_turu": butce.finans_turu,
                "yil": butce.yil,
                "ay": butce.ay,
                "hedef_tutar": _decimal_to_text(butce.hedef_tutar),
            }
            for butce in butceler
        ],
        "kategori_butceleri": [
            {
                "id": butce.id,
                "finans_turu": butce.finans_turu,
                "kategori_id": butce.kategori_id,
                "kategori_ad": butce.kategori.ad,
                "yil": butce.yil,
                "ay": butce.ay,
                "hedef_tutar": _decimal_to_text(butce.hedef_tutar),
            }
            for butce in kategori_butceleri
        ],
        "birikim_hedefleri": [
            {
                "id": hedef.id,
                "finans_turu": hedef.finans_turu,
                "hedef_adi": hedef.hedef_adi,
                "hedef_tutar": _decimal_to_text(hedef.hedef_tutar),
                "mevcut_tutar": _decimal_to_text(hedef.mevcut_tutar),
                "aylik_katki": _decimal_to_text(hedef.aylik_katki),
                "hedef_tarihi": _date_to_text(hedef.hedef_tarihi),
                "aktif": hedef.aktif,
                "aciklama": hedef.aciklama,
            }
            for hedef in birikim_hedefleri
        ],
        "tekrarlayan_odemeler": [
            {
                "id": odeme.id,
                "finans_turu": odeme.finans_turu,
                "kategori_id": odeme.kategori_id,
                "kategori_ad": odeme.kategori.ad,
                "odeme_adi": odeme.odeme_adi,
                "aciklama": odeme.aciklama,
                "tutar": _decimal_to_text(odeme.tutar),
                "baslangic_tarihi": _date_to_text(odeme.baslangic_tarihi),
                "tekrar_turu": odeme.tekrar_turu,
                "tekrar_araligi": odeme.tekrar_araligi,
                "aktif": odeme.aktif,
                "odeme_durumu": odeme.odeme_durumu,
                "son_olusturma_tarihi": _date_to_text(odeme.son_olusturma_tarihi),
                "son_gider_id": odeme.son_gider_id,
            }
            for odeme in odemeler
        ],
        "odeme_donemleri": [
            {
                "id": donem.id,
                "tekrarlayan_odeme_id": donem.tekrarlayan_odeme_id,
                "donem_yil": donem.donem_yil,
                "donem_ay": donem.donem_ay,
                "vade_tarihi": _date_to_text(donem.vade_tarihi),
                "durum": donem.durum,
            }
            for donem in donemler
        ],
    }


def _clear_user_backup_data(kullanici):
    OdemeDonemi.objects.filter(tekrarlayan_odeme__kullanici=kullanici).delete()
    TekrarlayanOdeme.objects.filter(kullanici=kullanici).delete()
    KategoriButcesi.objects.filter(kullanici=kullanici).delete()
    BirikimHedefi.objects.filter(kullanici=kullanici).delete()
    ButceHedefi.objects.filter(kullanici=kullanici).delete()
    Gelir.objects.filter(kullanici=kullanici).delete()
    Gider.objects.filter(kullanici=kullanici).delete()
    Kategori.objects.filter(kullanici=kullanici).delete()


def _restore_backup_payload(kullanici, payload, replace_existing=False):
    if not isinstance(payload, dict) or payload.get("meta", {}).get("app") != "FinansTakip":
        raise ValueError("Bu dosya geçerli bir FinansTakip yedeği değil.")

    if replace_existing:
        _clear_user_backup_data(kullanici)

    kategori_map = {}
    for item in payload.get("kategoriler", []):
        ad = str(item.get("ad", "")).strip()
        tur = item.get("tur")
        finans_turu = item.get("finans_turu") or FINANS_KISISEL
        if not ad or tur not in [Kategori.GELIR, Kategori.GIDER]:
            continue
        kategori, _ = Kategori.objects.get_or_create(
            kullanici=kullanici,
            finans_turu=finans_turu,
            ad=ad,
            tur=tur,
        )
        if item.get("id") is not None:
            kategori_map[str(item["id"])] = kategori

    gelir_sayisi = 0
    for item in payload.get("gelirler", []):
        tutar = _parse_decimal(item.get("tutar"))
        if not tutar or not item.get("tarih") or not item.get("aciklama"):
            continue
        Gelir.objects.create(
            kullanici=kullanici,
            finans_turu=item.get("finans_turu") or FINANS_KISISEL,
            tarih=item["tarih"],
            aciklama=str(item["aciklama"])[:200],
            tutar=tutar,
            kategori=str(item.get("kategori", ""))[:100],
        )
        gelir_sayisi += 1

    gider_map = {}
    gider_sayisi = 0
    for item in payload.get("giderler", []):
        tutar = _parse_decimal(item.get("tutar"))
        if not tutar or not item.get("tarih") or not item.get("aciklama"):
            continue
        gider = Gider.objects.create(
            kullanici=kullanici,
            finans_turu=item.get("finans_turu") or FINANS_KISISEL,
            tarih=item["tarih"],
            aciklama=str(item["aciklama"])[:200],
            tutar=tutar,
            kategori=str(item.get("kategori", ""))[:100],
        )
        if item.get("id") is not None:
            gider_map[str(item["id"])] = gider
        gider_sayisi += 1

    butce_sayisi = 0
    for item in payload.get("butce_hedefleri", []):
        hedef_tutar = _parse_decimal(item.get("hedef_tutar"))
        if not hedef_tutar or not item.get("yil") or not item.get("ay"):
            continue
        ButceHedefi.objects.update_or_create(
            kullanici=kullanici,
            finans_turu=item.get("finans_turu") or FINANS_KISISEL,
            yil=int(item["yil"]),
            ay=int(item["ay"]),
            defaults={"hedef_tutar": hedef_tutar},
        )
        butce_sayisi += 1

    kategori_butce_sayisi = 0
    for item in payload.get("kategori_butceleri", []):
        hedef_tutar = _parse_decimal(item.get("hedef_tutar"))
        kategori = kategori_map.get(str(item.get("kategori_id")))
        if not kategori and item.get("kategori_ad"):
            kategori = Kategori.objects.filter(
                kullanici=kullanici,
                finans_turu=item.get("finans_turu") or FINANS_KISISEL,
                ad=item.get("kategori_ad"),
                tur=Kategori.GIDER,
            ).first()
        if not hedef_tutar or not kategori or not item.get("yil") or not item.get("ay"):
            continue
        KategoriButcesi.objects.update_or_create(
            kullanici=kullanici,
            finans_turu=item.get("finans_turu") or FINANS_KISISEL,
            kategori=kategori,
            yil=int(item["yil"]),
            ay=int(item["ay"]),
            defaults={"hedef_tutar": hedef_tutar},
        )
        kategori_butce_sayisi += 1

    birikim_sayisi = 0
    for item in payload.get("birikim_hedefleri", []):
        hedef_tutar = _parse_decimal(item.get("hedef_tutar"))
        mevcut_tutar = _parse_decimal(item.get("mevcut_tutar")) or Decimal("0")
        aylik_katki = _parse_decimal(item.get("aylik_katki")) or Decimal("0")
        hedef_adi = str(item.get("hedef_adi", "")).strip()
        if not hedef_adi or not hedef_tutar:
            continue
        BirikimHedefi.objects.create(
            kullanici=kullanici,
            finans_turu=item.get("finans_turu") or FINANS_KISISEL,
            hedef_adi=hedef_adi[:120],
            hedef_tutar=hedef_tutar,
            mevcut_tutar=max(mevcut_tutar, Decimal("0")),
            aylik_katki=max(aylik_katki, Decimal("0")),
            hedef_tarihi=item.get("hedef_tarihi") or None,
            aktif=bool(item.get("aktif", True)),
            aciklama=str(item.get("aciklama", "")),
        )
        birikim_sayisi += 1

    odeme_map = {}
    odeme_sayisi = 0
    for item in payload.get("tekrarlayan_odemeler", []):
        tutar = _parse_decimal(item.get("tutar"))
        kategori = kategori_map.get(str(item.get("kategori_id")))
        if not kategori and item.get("kategori_ad"):
            kategori = Kategori.objects.filter(
                kullanici=kullanici,
                finans_turu=item.get("finans_turu") or FINANS_KISISEL,
                ad=item.get("kategori_ad"),
                tur=Kategori.GIDER,
            ).first()
        if not tutar or not kategori or not item.get("odeme_adi") or not item.get("baslangic_tarihi"):
            continue
        odeme = TekrarlayanOdeme.objects.create(
            kullanici=kullanici,
            finans_turu=item.get("finans_turu") or FINANS_KISISEL,
            kategori=kategori,
            odeme_adi=str(item["odeme_adi"])[:100],
            aciklama=str(item.get("aciklama", ""))[:200],
            tutar=tutar,
            baslangic_tarihi=item["baslangic_tarihi"],
            tekrar_turu=item.get("tekrar_turu") or TekrarlayanOdeme.AYLIK,
            tekrar_araligi=max(int(item.get("tekrar_araligi") or 1), 1),
            aktif=bool(item.get("aktif", True)),
            odeme_durumu=item.get("odeme_durumu") or TekrarlayanOdeme.BEKLIYOR,
            son_olusturma_tarihi=item.get("son_olusturma_tarihi") or None,
            son_gider=gider_map.get(str(item.get("son_gider_id"))),
        )
        if item.get("id") is not None:
            odeme_map[str(item["id"])] = odeme
        odeme_sayisi += 1

    donem_sayisi = 0
    for item in payload.get("odeme_donemleri", []):
        odeme = odeme_map.get(str(item.get("tekrarlayan_odeme_id")))
        if not odeme or not item.get("donem_yil") or not item.get("donem_ay") or not item.get("vade_tarihi"):
            continue
        OdemeDonemi.objects.update_or_create(
            tekrarlayan_odeme=odeme,
            donem_yil=int(item["donem_yil"]),
            donem_ay=int(item["donem_ay"]),
            defaults={
                "vade_tarihi": item["vade_tarihi"],
                "durum": item.get("durum") or OdemeDonemi.BEKLIYOR,
            },
        )
        donem_sayisi += 1

    return {
        "gelirler": gelir_sayisi,
        "giderler": gider_sayisi,
        "kategoriler": len(kategori_map),
        "butce_hedefleri": butce_sayisi,
        "kategori_butceleri": kategori_butce_sayisi,
        "birikim_hedefleri": birikim_sayisi,
        "tekrarlayan_odemeler": odeme_sayisi,
        "odeme_donemleri": donem_sayisi,
    }


def _append_rows(sheet, headers, rows):
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


@login_required
def yedekleme(request):
    context = {
        "gelir_sayisi": Gelir.objects.filter(kullanici=request.user).count(),
        "gider_sayisi": Gider.objects.filter(kullanici=request.user).count(),
        "kategori_sayisi": Kategori.objects.filter(kullanici=request.user).count(),
        "butce_sayisi": ButceHedefi.objects.filter(kullanici=request.user).count(),
        "kategori_butce_sayisi": KategoriButcesi.objects.filter(kullanici=request.user).count(),
        "birikim_hedefi_sayisi": BirikimHedefi.objects.filter(kullanici=request.user).count(),
        "tekrarlayan_odeme_sayisi": TekrarlayanOdeme.objects.filter(kullanici=request.user).count(),
        "odeme_donemi_sayisi": OdemeDonemi.objects.filter(tekrarlayan_odeme__kullanici=request.user).count(),
    }
    context.update(_finans_turu_context(FINANS_KISISEL))
    return render(request, "yedekleme.html", context)


@login_required
def yedekleme_json_indir(request):
    payload = _backup_payload(request.user)
    response = HttpResponse(
        json.dumps(payload, ensure_ascii=False, indent=2),
        content_type="application/json; charset=utf-8",
    )
    response["Content-Disposition"] = 'attachment; filename="finanstakip-yedek.json"'
    return response


@login_required
def yedekleme_excel_indir(request):
    payload = _backup_payload(request.user)
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheet = workbook.create_sheet("Gelirler")
    _append_rows(sheet, ["Tarih", "Finans Turu", "Aciklama", "Kategori", "Tutar"], [
        [item["tarih"], item["finans_turu"], item["aciklama"], item["kategori"], item["tutar"]]
        for item in payload["gelirler"]
    ])

    sheet = workbook.create_sheet("Giderler")
    _append_rows(sheet, ["Tarih", "Finans Turu", "Aciklama", "Kategori", "Tutar"], [
        [item["tarih"], item["finans_turu"], item["aciklama"], item["kategori"], item["tutar"]]
        for item in payload["giderler"]
    ])

    sheet = workbook.create_sheet("Kategoriler")
    _append_rows(sheet, ["Finans Turu", "Ad", "Tur"], [
        [item["finans_turu"], item["ad"], item["tur"]]
        for item in payload["kategoriler"]
    ])

    sheet = workbook.create_sheet("Butce Hedefleri")
    _append_rows(sheet, ["Finans Turu", "Yil", "Ay", "Hedef Tutar"], [
        [item["finans_turu"], item["yil"], item["ay"], item["hedef_tutar"]]
        for item in payload["butce_hedefleri"]
    ])

    sheet = workbook.create_sheet("Kategori Butceleri")
    _append_rows(sheet, ["Finans Turu", "Kategori", "Yil", "Ay", "Hedef Tutar"], [
        [item["finans_turu"], item["kategori_ad"], item["yil"], item["ay"], item["hedef_tutar"]]
        for item in payload["kategori_butceleri"]
    ])

    sheet = workbook.create_sheet("Birikim Hedefleri")
    _append_rows(sheet, ["Finans Turu", "Hedef", "Hedef Tutar", "Mevcut Tutar", "Aylik Katki", "Hedef Tarihi", "Aktif", "Aciklama"], [
        [
            item["finans_turu"],
            item["hedef_adi"],
            item["hedef_tutar"],
            item["mevcut_tutar"],
            item["aylik_katki"],
            item["hedef_tarihi"],
            item["aktif"],
            item["aciklama"],
        ]
        for item in payload["birikim_hedefleri"]
    ])

    sheet = workbook.create_sheet("Tekrarlayan Odemeler")
    _append_rows(sheet, ["Odeme", "Finans Turu", "Kategori", "Tutar", "Baslangic", "Tekrar", "Aralik", "Aktif", "Durum"], [
        [
            item["odeme_adi"],
            item["finans_turu"],
            item["kategori_ad"],
            item["tutar"],
            item["baslangic_tarihi"],
            item["tekrar_turu"],
            item["tekrar_araligi"],
            item["aktif"],
            item["odeme_durumu"],
        ]
        for item in payload["tekrarlayan_odemeler"]
    ])

    sheet = workbook.create_sheet("Odeme Donemleri")
    _append_rows(sheet, ["Tekrarlayan Odeme ID", "Yil", "Ay", "Vade", "Durum"], [
        [
            item["tekrarlayan_odeme_id"],
            item["donem_yil"],
            item["donem_ay"],
            item["vade_tarihi"],
            item["durum"],
        ]
        for item in payload["odeme_donemleri"]
    ])

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="finanstakip-yedek.xlsx"'
    buffer.close()
    return response


@login_required
def yedekleme_geri_yukle(request):
    if request.method != "POST":
        return redirect("yedekleme")

    dosya = request.FILES.get("yedek_dosyasi")
    if not dosya:
        messages.error(request, "Lütfen bir JSON yedek dosyası seçin.")
        return redirect("yedekleme")

    try:
        payload = json.loads(dosya.read().decode("utf-8"))
        with transaction.atomic():
            sonuc = _restore_backup_payload(
                request.user,
                payload,
                replace_existing=bool(request.POST.get("replace_existing")),
            )
    except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as hata:
        messages.error(request, str(hata))
    else:
        messages.success(
            request,
            "Yedek geri yüklendi: "
            f"{sonuc['gelirler']} gelir, {sonuc['giderler']} gider, "
            f"{sonuc['kategoriler']} kategori, {sonuc['butce_hedefleri']} bütçe hedefi, "
            f"{sonuc['kategori_butceleri']} kategori bütçesi, {sonuc['birikim_hedefleri']} birikim hedefi, "
            f"{sonuc['tekrarlayan_odemeler']} tekrarlayan ödeme ve {sonuc['odeme_donemleri']} ödeme dönemi işlendi.",
        )

    return redirect("yedekleme")


def _pdf_font_adi():
    font_adi = "TurkceFont"
    if font_adi in pdfmetrics.getRegisteredFontNames():
        return font_adi

    font_yollari = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/Library/Fonts/Arial Unicode.ttf"),
    ]

    for font_yolu in font_yollari:
        if font_yolu.exists():
            pdfmetrics.registerFont(TTFont(font_adi, str(font_yolu)))
            return font_adi

    return "Helvetica"


@never_cache
@ensure_csrf_cookie
def kayit(request):
    if request.user.is_authenticated:
        return redirect("home")

    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            kullanici = form.save()
            login(request, kullanici)
            return redirect("home")
    else:
        form = UserCreationForm()

    return render(request, "kayit.html", {"form": form})


@never_cache
@ensure_csrf_cookie
def giris(request):
    if request.user.is_authenticated:
        return redirect("home")

    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            kullanici = form.get_user()
            login(request, kullanici)
            _tekrarlayan_odemeleri_olustur(kullanici)
            return redirect("home")
    else:
        form = AuthenticationForm()

    return render(request, "giris.html", {"form": form})


@login_required
@never_cache
def cikis(request):
    logout(request)
    return redirect("giris")


@login_required
def home(request):
    finans_turu = _secilen_finans_turu(request.GET)

    toplam_gelir = Gelir.objects.filter(kullanici=request.user, finans_turu=finans_turu).aggregate(
        Sum("tutar")
    )["tutar__sum"] or 0

    toplam_gider = Gider.objects.filter(kullanici=request.user, finans_turu=finans_turu).aggregate(
        Sum("tutar")
    )["tutar__sum"] or 0

    bakiye = toplam_gelir - toplam_gider

    son_gelirler = Gelir.objects.filter(kullanici=request.user, finans_turu=finans_turu).order_by("-id")[:5]
    son_giderler = Gider.objects.filter(kullanici=request.user, finans_turu=finans_turu).order_by("-id")[:5]
    hizli_gider_kategorileri = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GIDER,
    ).order_by("ad")

    context = {
        "toplam_gelir": toplam_gelir,
        "toplam_gider": toplam_gider,
        "bakiye": bakiye,
        "son_gelirler": son_gelirler,
        "son_giderler": son_giderler,
        "hizli_gider_kategorileri": hizli_gider_kategorileri,
        "hizli_gider_acik": request.GET.get("quick_add") == "1",
    }
    butce_verileri = _aylik_butce_verileri(request.user, finans_turu)
    odeme_verileri = _tekrarlayan_odeme_verileri(request.user, finans_turu)
    kategori_butce_verileri = _kategori_butce_verileri(request.user, finans_turu)
    birikim_hedefi_verileri = _birikim_hedefi_verileri(request.user, finans_turu)
    grafik_verileri = _dashboard_grafik_verileri(
        request.user,
        finans_turu,
        butce_verileri,
        odeme_verileri,
        kategori_butce_verileri,
    )
    context.update(butce_verileri)
    context.update(odeme_verileri)
    context.update(kategori_butce_verileri)
    context.update(birikim_hedefi_verileri)
    context.update(grafik_verileri)
    context.update(_finansal_tavsiyeler(grafik_verileri, butce_verileri, odeme_verileri, kategori_butce_verileri))

    return render(request, "home.html", context)


@login_required
@never_cache
def hizli_gider_ekle(request):
    if request.method != "POST":
        return redirect("home")

    finans_turu = _secilen_finans_turu(request.POST)
    tutar = _parse_decimal(request.POST.get("tutar"))
    kategori = Kategori.objects.filter(
        id=request.POST.get("kategori"),
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GIDER,
    ).first()
    aciklama = (request.POST.get("aciklama") or "").strip()

    if tutar is None or tutar <= 0:
        messages.error(request, "Hızlı harcama için geçerli bir tutar girin.")
        return redirect(f"/?finans_turu={finans_turu}&quick_add=1")

    if not kategori:
        messages.error(request, "Seçilen finans türü için geçerli bir gider kategorisi seçin.")
        return redirect(f"/?finans_turu={finans_turu}&quick_add=1")

    Gider.objects.create(
        kullanici=request.user,
        finans_turu=finans_turu,
        tarih=timezone.now().date(),
        aciklama=aciklama or "Hızlı harcama",
        tutar=tutar,
        kategori=kategori.ad,
    )
    messages.success(request, "Hızlı harcama başarıyla eklendi.")
    return redirect(f"/?finans_turu={finans_turu}")


@login_required
def butce_hedefi(request):
    finans_turu = _secilen_finans_turu(request.POST if request.method == "POST" else request.GET)
    butce_verileri = _aylik_butce_verileri(request.user, finans_turu)
    hedef = butce_verileri["butce_hedefi"]
    hata = None

    if request.method == "POST":
        hedef_tutar = request.POST.get("hedef_tutar")

        try:
            hedef_tutar = Decimal(hedef_tutar)
        except (InvalidOperation, TypeError):
            hata = "Lütfen geçerli bir bütçe hedefi girin."
        else:
            if hedef_tutar <= 0:
                hata = "Bütçe hedefi sıfırdan büyük olmalıdır."
            else:
                hedef, _ = ButceHedefi.objects.update_or_create(
                    kullanici=request.user,
                    finans_turu=finans_turu,
                    yil=butce_verileri["butce_yil"],
                    ay=butce_verileri["butce_ay"],
                    defaults={"hedef_tutar": hedef_tutar},
                )
                return redirect(f"/?finans_turu={finans_turu}")

    return render(request, "butce_hedefi.html", {
        **butce_verileri,
        "hedef": hedef,
        "hata": hata,
    })


@login_required
def kategori_butceleri(request):
    finans_turu = _secilen_finans_turu(request.POST if request.method == "POST" else request.GET)
    bugun = timezone.now().date()
    yil = _parse_int(request.POST.get("yil") or request.GET.get("yil"), bugun.year)
    ay = _parse_int(request.POST.get("ay") or request.GET.get("ay"), bugun.month)
    kategoriler = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GIDER,
    ).order_by("ad")
    hata = None

    if request.method == "POST":
        kategori = kategoriler.filter(id=request.POST.get("kategori")).first()
        hedef_tutar = _parse_decimal(request.POST.get("hedef_tutar"))

        if not kategori:
            hata = "Kategori bütçesi için seçili finans türüne ait gider kategorisi seçmelisin."
        elif ay < 1 or ay > 12:
            hata = "Ay değeri 1 ile 12 arasında olmalıdır."
        elif not hedef_tutar or hedef_tutar <= 0:
            hata = "Hedef tutar sıfırdan büyük olmalıdır."
        else:
            KategoriButcesi.objects.update_or_create(
                kullanici=request.user,
                finans_turu=finans_turu,
                kategori=kategori,
                yil=yil,
                ay=ay,
                defaults={"hedef_tutar": hedef_tutar},
            )
            messages.success(request, "Kategori bütçesi kaydedildi.")
            return redirect(f"/kategori-butceleri/?finans_turu={finans_turu}&yil={yil}&ay={ay}")

    butceler = KategoriButcesi.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        yil=yil,
        ay=ay,
    ).select_related("kategori").order_by("kategori__ad")

    kayitlar = []
    for butce in butceler:
        harcanan = Gider.objects.filter(
            kullanici=request.user,
            finans_turu=finans_turu,
            kategori=butce.kategori.ad,
            tarih__year=yil,
            tarih__month=ay,
        ).aggregate(Sum("tutar"))["tutar__sum"] or Decimal("0")
        oran = (harcanan / butce.hedef_tutar) * 100 if butce.hedef_tutar else Decimal("0")
        oran_float = round(float(oran), 2)
        renk = "danger" if harcanan > butce.hedef_tutar else "warning" if oran_float >= 80 else "success"
        kayitlar.append({
            "butce": butce,
            "harcanan": harcanan,
            "kalan": butce.hedef_tutar - harcanan,
            "oran": oran_float,
            "progress_orani": min(oran_float, 100),
            "renk": renk,
        })

    return render(request, "kategori_butceleri.html", {
        "kategoriler": kategoriler,
        "butce_kayitlari": kayitlar,
        "yil": yil,
        "ay": ay,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })


@login_required
def kategori_butcesi_duzenle(request, id):
    butce = get_object_or_404(KategoriButcesi, id=id, kullanici=request.user)
    finans_turu = _secilen_finans_turu(request.POST) if request.method == "POST" else butce.finans_turu
    kategoriler = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GIDER,
    ).order_by("ad")
    hata = None

    if request.method == "POST":
        kategori = kategoriler.filter(id=request.POST.get("kategori")).first()
        hedef_tutar = _parse_decimal(request.POST.get("hedef_tutar"))
        yil = _parse_int(request.POST.get("yil"), butce.yil)
        ay = _parse_int(request.POST.get("ay"), butce.ay)

        if not kategori:
            hata = "Seçili finans türüne ait gider kategorisi seçmelisin."
        elif ay < 1 or ay > 12:
            hata = "Ay değeri 1 ile 12 arasında olmalıdır."
        elif not hedef_tutar or hedef_tutar <= 0:
            hata = "Hedef tutar sıfırdan büyük olmalıdır."
        elif KategoriButcesi.objects.filter(
            kullanici=request.user,
            finans_turu=finans_turu,
            kategori=kategori,
            yil=yil,
            ay=ay,
        ).exclude(id=butce.id).exists():
            hata = "Bu kategori için aynı dönemde zaten bütçe var."
        else:
            butce.finans_turu = finans_turu
            butce.kategori = kategori
            butce.yil = yil
            butce.ay = ay
            butce.hedef_tutar = hedef_tutar
            butce.save()
            messages.success(request, "Kategori bütçesi güncellendi.")
            return redirect(f"/kategori-butceleri/?finans_turu={finans_turu}&yil={yil}&ay={ay}")

    return render(request, "kategori_butcesi_duzenle.html", {
        "butce": butce,
        "kategoriler": kategoriler,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })


@login_required
def kategori_butcesi_sil(request, id):
    butce = get_object_or_404(KategoriButcesi, id=id, kullanici=request.user)
    finans_turu = butce.finans_turu
    yil = butce.yil
    ay = butce.ay
    butce.delete()
    messages.success(request, "Kategori bütçesi silindi.")
    return redirect(f"/kategori-butceleri/?finans_turu={finans_turu}&yil={yil}&ay={ay}")


@login_required
def birikim_hedefleri(request):
    finans_turu = _secilen_finans_turu(request.POST if request.method == "POST" else request.GET)
    hata = None

    if request.method == "POST":
        hedef_tutar = _parse_decimal(request.POST.get("hedef_tutar"))
        mevcut_tutar = _parse_decimal(request.POST.get("mevcut_tutar")) or Decimal("0")
        aylik_katki = _parse_decimal(request.POST.get("aylik_katki")) or Decimal("0")
        hedef_adi = request.POST.get("hedef_adi", "").strip()

        if not hedef_adi:
            hata = "Hedef adı zorunludur."
        elif not hedef_tutar or hedef_tutar <= 0:
            hata = "Hedef tutar sıfırdan büyük olmalıdır."
        elif mevcut_tutar < 0 or aylik_katki < 0:
            hata = "Mevcut tutar ve aylık katkı negatif olamaz."
        else:
            BirikimHedefi.objects.create(
                kullanici=request.user,
                finans_turu=finans_turu,
                hedef_adi=hedef_adi,
                hedef_tutar=hedef_tutar,
                mevcut_tutar=mevcut_tutar,
                aylik_katki=aylik_katki,
                hedef_tarihi=request.POST.get("hedef_tarihi") or None,
                aktif=bool(request.POST.get("aktif")),
                aciklama=request.POST.get("aciklama", "").strip(),
            )
            messages.success(request, "Birikim hedefi oluşturuldu.")
            return redirect(f"/birikim-hedefleri/?finans_turu={finans_turu}")

    hedefler = _birikim_hedefi_verileri(request.user, finans_turu)["birikim_hedefi_kayitlari"]
    tum_hedefler = BirikimHedefi.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
    ).order_by("-aktif", "hedef_tarihi", "hedef_adi")

    return render(request, "birikim_hedefleri.html", {
        "hedef_kayitlari": hedefler,
        "tum_hedefler": tum_hedefler,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })


@login_required
def birikim_hedefi_duzenle(request, id):
    hedef = get_object_or_404(BirikimHedefi, id=id, kullanici=request.user)
    finans_turu = _secilen_finans_turu(request.POST) if request.method == "POST" else hedef.finans_turu
    hata = None

    if request.method == "POST":
        hedef_tutar = _parse_decimal(request.POST.get("hedef_tutar"))
        mevcut_tutar = _parse_decimal(request.POST.get("mevcut_tutar")) or Decimal("0")
        aylik_katki = _parse_decimal(request.POST.get("aylik_katki")) or Decimal("0")
        hedef_adi = request.POST.get("hedef_adi", "").strip()

        if not hedef_adi:
            hata = "Hedef adı zorunludur."
        elif not hedef_tutar or hedef_tutar <= 0:
            hata = "Hedef tutar sıfırdan büyük olmalıdır."
        elif mevcut_tutar < 0 or aylik_katki < 0:
            hata = "Mevcut tutar ve aylık katkı negatif olamaz."
        else:
            hedef.finans_turu = finans_turu
            hedef.hedef_adi = hedef_adi
            hedef.hedef_tutar = hedef_tutar
            hedef.mevcut_tutar = mevcut_tutar
            hedef.aylik_katki = aylik_katki
            hedef.hedef_tarihi = request.POST.get("hedef_tarihi") or None
            hedef.aktif = bool(request.POST.get("aktif"))
            hedef.aciklama = request.POST.get("aciklama", "").strip()
            hedef.save()
            messages.success(request, "Birikim hedefi güncellendi.")
            return redirect(f"/birikim-hedefleri/?finans_turu={finans_turu}")

    return render(request, "birikim_hedefi_duzenle.html", {
        "hedef": hedef,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })


@login_required
def birikim_hedefi_sil(request, id):
    hedef = get_object_or_404(BirikimHedefi, id=id, kullanici=request.user)
    finans_turu = hedef.finans_turu
    hedef.delete()
    messages.success(request, "Birikim hedefi silindi.")
    return redirect(f"/birikim-hedefleri/?finans_turu={finans_turu}")


@login_required
def birikim_hedefi_katki(request, id):
    hedef = get_object_or_404(BirikimHedefi, id=id, kullanici=request.user)
    if request.method == "POST":
        tutar = _parse_decimal(request.POST.get("katki_tutar"))
        if not tutar or tutar <= 0:
            messages.error(request, "Katkı tutarı sıfırdan büyük olmalıdır.")
        else:
            hedef.mevcut_tutar += tutar
            hedef.save(update_fields=["mevcut_tutar"])
            messages.success(request, "Katkı hedefe eklendi.")
    return redirect(f"/birikim-hedefleri/?finans_turu={hedef.finans_turu}")


@login_required
def kategoriler(request):
    finans_turu = _secilen_finans_turu(request.POST if request.method == "POST" else request.GET)
    hata = None

    if request.method == "POST":
        ad = request.POST.get("ad", "").strip()
        tur = request.POST.get("tur")

        if not ad or tur not in [Kategori.GELIR, Kategori.GIDER]:
            hata = "Lütfen kategori adı ve türünü doğru girin."
        elif Kategori.objects.filter(kullanici=request.user, finans_turu=finans_turu, ad=ad, tur=tur).exists():
            hata = "Bu kategori zaten mevcut."
        else:
            Kategori.objects.create(kullanici=request.user, finans_turu=finans_turu, ad=ad, tur=tur)
            return redirect(f"/kategoriler/?finans_turu={finans_turu}")

    kullanici_kategorileri = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
    ).order_by("tur", "ad")
    return render(request, "kategoriler.html", {
        "kategoriler": kullanici_kategorileri,
        "turler": Kategori.TUR_SECENEKLERI,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })


@login_required
def kategori_duzenle(request, id):
    kategori = get_object_or_404(Kategori, id=id, kullanici=request.user)
    finans_turu = kategori.finans_turu
    hata = None

    if request.method == "POST":
        ad = request.POST.get("ad", "").strip()
        tur = request.POST.get("tur")
        finans_turu = _secilen_finans_turu(request.POST)

        if not ad or tur not in [Kategori.GELIR, Kategori.GIDER]:
            hata = "Lütfen kategori adı ve türünü doğru girin."
        elif Kategori.objects.filter(kullanici=request.user, finans_turu=finans_turu, ad=ad, tur=tur).exclude(id=kategori.id).exists():
            hata = "Bu kategori zaten mevcut."
        else:
            kategori.ad = ad
            kategori.tur = tur
            kategori.finans_turu = finans_turu
            kategori.save()
            return redirect(f"/kategoriler/?finans_turu={finans_turu}")

    return render(request, "kategori_duzenle.html", {
        "kategori": kategori,
        "turler": Kategori.TUR_SECENEKLERI,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })


@login_required
def kategori_sil(request, id):
    kategori = get_object_or_404(Kategori, id=id, kullanici=request.user)
    finans_turu = kategori.finans_turu
    kategori.delete()
    return redirect(f"/kategoriler/?finans_turu={finans_turu}")


@login_required
def tekrarlayan_odemeler(request):
    finans_turu = _secilen_finans_turu(request.POST if request.method == "POST" else request.GET)
    kategoriler = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GIDER,
    ).order_by("ad")
    hata = None

    if request.method == "POST":
        kategori_id = request.POST.get("kategori")
        tekrar_turu = request.POST.get("tekrar_turu")
        kategori = None
        if kategori_id:
            kategori = kategoriler.filter(id=kategori_id).first()

        try:
            tutar = Decimal(request.POST.get("tutar"))
        except (InvalidOperation, TypeError):
            tutar = None

        try:
            tekrar_araligi = int(request.POST.get("tekrar_araligi"))
        except (TypeError, ValueError):
            tekrar_araligi = 0

        if not kategori:
            hata = "Tekrarlayan ödeme için seçili finans türüne ait gider kategorisi seçmelisin."
        elif tekrar_turu not in [deger for deger, _ in TekrarlayanOdeme.TEKRAR_TURU_SECENEKLERI]:
            hata = "Lütfen geçerli bir tekrar türü seçin."
        elif tekrar_araligi <= 0:
            hata = "Tekrar aralığı pozitif tam sayı olmalıdır."
        elif not request.POST.get("odeme_adi", "").strip():
            hata = "Ödeme adı zorunludur."
        elif tutar is None or tutar <= 0:
            hata = "Tutar sıfırdan büyük olmalıdır."
        else:
            aktif = bool(request.POST.get("aktif"))
            TekrarlayanOdeme.objects.create(
                kullanici=request.user,
                finans_turu=finans_turu,
                kategori=kategori,
                odeme_adi=request.POST["odeme_adi"].strip(),
                aciklama=request.POST.get("aciklama", "").strip(),
                tutar=tutar,
                baslangic_tarihi=request.POST["baslangic_tarihi"],
                tekrar_turu=tekrar_turu,
                tekrar_araligi=tekrar_araligi,
                aktif=aktif,
                odeme_durumu=TekrarlayanOdeme.BEKLIYOR if aktif else TekrarlayanOdeme.IPTAL,
            )
            return redirect(f"/tekrarlayan-odemeler/?finans_turu={finans_turu}")

    odemeler = TekrarlayanOdeme.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
    ).select_related("kategori").order_by("-aktif", "baslangic_tarihi", "odeme_adi")
    bugun = timezone.now().date()
    _odeme_donemlerini_olustur(request.user, finans_turu)
    odeme_kayitlari = []
    for odeme in odemeler:
        vade_tarihi = _bekleyen_odeme_tarihi(odeme)
        odeme_kayitlari.append({
            "odeme": odeme,
            "vade_tarihi": vade_tarihi,
            "durum": _odeme_durumu_verisi(odeme, vade_tarihi, bugun),
        })
    donem_kayitlari = []
    donemler = OdemeDonemi.objects.filter(
        tekrarlayan_odeme__kullanici=request.user,
        tekrarlayan_odeme__finans_turu=finans_turu,
    ).select_related("tekrarlayan_odeme", "tekrarlayan_odeme__kategori").order_by("-donem_yil", "-donem_ay", "vade_tarihi")
    for donem in donemler:
        donem_kayitlari.append({
            "donem": donem,
            "odeme": donem.tekrarlayan_odeme,
            "durum": _donem_durumu_guncelle(donem, bugun),
        })

    return render(request, "tekrarlayan_odemeler.html", {
        "odeme_kayitlari": odeme_kayitlari,
        "donem_kayitlari": donem_kayitlari,
        "kategoriler": kategoriler,
        "tekrar_turleri": TekrarlayanOdeme.TEKRAR_TURU_SECENEKLERI,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })


@login_required
def tekrarlayan_odeme_duzenle(request, id):
    odeme = get_object_or_404(TekrarlayanOdeme, id=id, kullanici=request.user)
    finans_turu = _secilen_finans_turu(request.POST) if request.method == "POST" else odeme.finans_turu
    kategoriler = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GIDER,
    ).order_by("ad")
    hata = None

    if request.method == "POST":
        kategori_id = request.POST.get("kategori")
        tekrar_turu = request.POST.get("tekrar_turu")
        kategori = None
        if kategori_id:
            kategori = kategoriler.filter(id=kategori_id).first()

        try:
            tutar = Decimal(request.POST.get("tutar"))
        except (InvalidOperation, TypeError):
            tutar = None

        try:
            tekrar_araligi = int(request.POST.get("tekrar_araligi"))
        except (TypeError, ValueError):
            tekrar_araligi = 0

        if not kategori:
            hata = "Tekrarlayan ödeme için seçili finans türüne ait gider kategorisi seçmelisin."
        elif tekrar_turu not in [deger for deger, _ in TekrarlayanOdeme.TEKRAR_TURU_SECENEKLERI]:
            hata = "Lütfen geçerli bir tekrar türü seçin."
        elif tekrar_araligi <= 0:
            hata = "Tekrar aralığı pozitif tam sayı olmalıdır."
        elif not request.POST.get("odeme_adi", "").strip():
            hata = "Ödeme adı zorunludur."
        elif tutar is None or tutar <= 0:
            hata = "Tutar sıfırdan büyük olmalıdır."
        else:
            aktif = bool(request.POST.get("aktif"))
            odeme.finans_turu = finans_turu
            odeme.kategori = kategori
            odeme.odeme_adi = request.POST["odeme_adi"].strip()
            odeme.aciklama = request.POST.get("aciklama", "").strip()
            odeme.tutar = tutar
            odeme.baslangic_tarihi = request.POST["baslangic_tarihi"]
            odeme.tekrar_turu = tekrar_turu
            odeme.tekrar_araligi = tekrar_araligi
            odeme.aktif = aktif
            if not aktif:
                odeme.odeme_durumu = TekrarlayanOdeme.IPTAL
            elif odeme.odeme_durumu == TekrarlayanOdeme.IPTAL:
                odeme.odeme_durumu = TekrarlayanOdeme.BEKLIYOR
            odeme.save()
            return redirect(f"/tekrarlayan-odemeler/?finans_turu={finans_turu}")

    return render(request, "tekrarlayan_odeme_duzenle.html", {
        "odeme": odeme,
        "kategoriler": kategoriler,
        "tekrar_turleri": TekrarlayanOdeme.TEKRAR_TURU_SECENEKLERI,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })


@login_required
def tekrarlayan_odeme_sil(request, id):
    odeme = get_object_or_404(TekrarlayanOdeme, id=id, kullanici=request.user)
    finans_turu = odeme.finans_turu
    odeme.delete()
    return redirect(f"/tekrarlayan-odemeler/?finans_turu={finans_turu}")


@login_required
def tekrarlayan_odeme_odendi(request, id):
    odeme = get_object_or_404(TekrarlayanOdeme, id=id, kullanici=request.user, aktif=True)
    finans_turu = odeme.finans_turu
    vade_tarihi = _bekleyen_odeme_tarihi(odeme)
    _tekrarlayan_odeme_gideri_olustur(odeme, vade_tarihi)
    return redirect(f"/tekrarlayan-odemeler/?finans_turu={finans_turu}")


@login_required
def odeme_donemi_odendi(request, id):
    donem = get_object_or_404(
        OdemeDonemi,
        id=id,
        tekrarlayan_odeme__kullanici=request.user,
    )
    finans_turu = donem.tekrarlayan_odeme.finans_turu

    if donem.durum != OdemeDonemi.ODENDI:
        _odeme_donemi_gideri_olustur(donem)

    return redirect(f"/tekrarlayan-odemeler/?finans_turu={finans_turu}")

@login_required
def gelir_ekle(request):
    finans_turu = _secilen_finans_turu(request.POST if request.method == "POST" else request.GET)
    gelir_kategorileri = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GELIR,
    ).order_by("ad")
    hata = None

    if request.method == "POST":
        kategori_id = request.POST.get("kategori")
        kategori = None
        if kategori_id:
            kategori = Kategori.objects.filter(
                id=kategori_id,
                kullanici=request.user,
                finans_turu=finans_turu,
                tur=Kategori.GELIR,
            ).first()

        if not kategori:
            hata = "Gelir kaydı için önce gelir kategorisi seçmelisin."
        else:
            Gelir.objects.create(
                kullanici=request.user,
                finans_turu=finans_turu,
                tarih=request.POST["tarih"],
                aciklama=request.POST["aciklama"],
                tutar=request.POST["tutar"],
                kategori=kategori.ad
            )
            return redirect(f"/gelir-ekle/?finans_turu={finans_turu}")

    gelirler = Gelir.objects.filter(kullanici=request.user, finans_turu=finans_turu).order_by("-tarih")

    baslangic = request.GET.get("baslangic")
    bitis = request.GET.get("bitis")
    kategori = request.GET.get("kategori")

    if baslangic:
        gelirler = gelirler.filter(tarih__gte=baslangic)

    if bitis:
        gelirler = gelirler.filter(tarih__lte=bitis)
    if kategori:
        gelirler = gelirler.filter(kategori=kategori)
    
    return render(request, "gelir_ekle.html", {
        "gelirler": gelirler,
        "gelir_kategorileri": gelir_kategorileri,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })


@login_required
def gider_ekle(request):
    finans_turu = _secilen_finans_turu(request.POST if request.method == "POST" else request.GET)
    gider_kategorileri = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GIDER,
    ).order_by("ad")
    hata = None

    if request.method == "POST":
        kategori_id = request.POST.get("kategori")
        kategori = None
        if kategori_id:
            kategori = Kategori.objects.filter(
                id=kategori_id,
                kullanici=request.user,
                finans_turu=finans_turu,
                tur=Kategori.GIDER,
            ).first()

        if not kategori:
            hata = "Gider kaydı için önce gider kategorisi seçmelisin."
        else:
            Gider.objects.create(
                kullanici=request.user,
                finans_turu=finans_turu,
                tarih=request.POST["tarih"],
                aciklama=request.POST["aciklama"],
                tutar=request.POST["tutar"],
                kategori=kategori.ad
            )
            return redirect(f"/gider-ekle/?finans_turu={finans_turu}")

    giderler = Gider.objects.filter(kullanici=request.user, finans_turu=finans_turu).order_by("-tarih")

    baslangic = request.GET.get("baslangic")
    bitis = request.GET.get("bitis")
    kategori = request.GET.get("kategori")

    if baslangic:
        giderler = giderler.filter(tarih__gte=baslangic)

    if bitis:
        giderler = giderler.filter(tarih__lte=bitis)

    if kategori:
        giderler = giderler.filter(kategori=kategori)

    return render(request, "gider_ekle.html", {
        "giderler": giderler,
        "gider_kategorileri": gider_kategorileri,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })

@login_required
def raporlar(request):
    ay = request.GET.get("ay")
    finans_turu = _secilen_finans_turu(request.GET)
    return render(request, "raporlar.html", _rapor_verileri(request.user, finans_turu, ay))


@login_required
def rapor_pdf(request):
    ay = request.GET.get("ay")
    finans_turu = _secilen_finans_turu(request.GET)
    veriler = _rapor_verileri(request.user, finans_turu, ay)
    buffer = BytesIO()
    font_adi = _pdf_font_adi()

    dosya_adi = f"finans-raporu-{finans_turu}-{ay}.pdf" if ay else f"finans-raporu-{finans_turu}.pdf"
    belge = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title="Finans Raporu",
    )

    stiller = getSampleStyleSheet()
    baslik_stili = ParagraphStyle(
        "Baslik",
        parent=stiller["Title"],
        fontName=font_adi,
        fontSize=20,
        alignment=TA_CENTER,
        spaceAfter=18,
    )
    normal_stili = ParagraphStyle(
        "NormalTurkce",
        parent=stiller["Normal"],
        fontName=font_adi,
        fontSize=10,
        leading=14,
    )
    sag_stili = ParagraphStyle(
        "SagTurkce",
        parent=normal_stili,
        alignment=TA_RIGHT,
    )

    pdf_icerigi = [
        Paragraph("Finans Raporu", baslik_stili),
        Paragraph(f"Rapor dönemi: {ay or 'Tüm kayıtlar'}", normal_stili),
        Spacer(1, 12),
    ]

    ozet_tablosu = Table([
        ["Başlık", "Tutar"],
        ["Toplam Gelir", f"{veriler['toplam_gelir']} TL"],
        ["Toplam Gider", f"{veriler['toplam_gider']} TL"],
        ["Kalan Bakiye", f"{veriler['bakiye']} TL"],
        ["Aylık Gelir", f"{veriler['aylik_gelir']} TL"],
        ["Aylık Gider", f"{veriler['aylik_gider']} TL"],
        ["Aylık Bakiye", f"{veriler['aylik_bakiye']} TL"],
    ], colWidths=[8 * cm, 6 * cm])
    ozet_tablosu.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_adi),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#343a40")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    pdf_icerigi.append(ozet_tablosu)
    pdf_icerigi.append(Spacer(1, 18))
    pdf_icerigi.append(Paragraph("Kategori Bazlı Giderler", normal_stili))
    pdf_icerigi.append(Spacer(1, 8))

    kategori_satirlari = [["Kategori", "Tutar"]]
    for kategori in veriler["kategori_detaylari"]:
        kategori_satirlari.append([
            Paragraph(escape(str(kategori["ad"])), normal_stili),
            Paragraph(f"{kategori['tutar']} TL", sag_stili),
        ])

    if len(kategori_satirlari) == 1:
        kategori_satirlari.append(["Kayıt bulunamadı", "0 TL"])

    kategori_tablosu = Table(kategori_satirlari, colWidths=[8 * cm, 6 * cm])
    kategori_tablosu.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_adi),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    pdf_icerigi.append(kategori_tablosu)

    belge.build(pdf_icerigi)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{dosya_adi}"'
    buffer.close()
    return response

@login_required
def gelir_sil(request, id):
    gelir = get_object_or_404(Gelir, id=id, kullanici=request.user)
    finans_turu = gelir.finans_turu
    gelir.delete()
    return redirect(f"/gelir-ekle/?finans_turu={finans_turu}")


@login_required
def gider_sil(request, id):
    gider = get_object_or_404(Gider, id=id, kullanici=request.user)
    finans_turu = gider.finans_turu
    gider.delete()
    return redirect(f"/gider-ekle/?finans_turu={finans_turu}")

@login_required
def gider_duzenle(request, id):
    gider = get_object_or_404(Gider, id=id, kullanici=request.user)
    finans_turu = _secilen_finans_turu(request.POST) if request.method == "POST" else gider.finans_turu
    gider_kategorileri = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GIDER,
    ).order_by("ad")
    hata = None

    if request.method == "POST":
        kategori_id = request.POST.get("kategori")
        kategori = None
        if kategori_id:
            kategori = Kategori.objects.filter(
                id=kategori_id,
                kullanici=request.user,
                finans_turu=finans_turu,
                tur=Kategori.GIDER,
            ).first()

        if not kategori:
            hata = "Gider kaydı için gider kategorisi seçmelisin."
        else:
            gider.tarih = request.POST["tarih"]
            gider.aciklama = request.POST["aciklama"]
            gider.tutar = request.POST["tutar"]
            gider.finans_turu = finans_turu
            gider.kategori = kategori.ad
            gider.save()

            return redirect(f"/gider-ekle/?finans_turu={finans_turu}")

    return render(request, "gider_duzenle.html", {
        "gider": gider,
        "gider_kategorileri": gider_kategorileri,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })

@login_required
def gelir_duzenle(request, id):
    gelir = get_object_or_404(Gelir, id=id, kullanici=request.user)
    finans_turu = _secilen_finans_turu(request.POST) if request.method == "POST" else gelir.finans_turu
    gelir_kategorileri = Kategori.objects.filter(
        kullanici=request.user,
        finans_turu=finans_turu,
        tur=Kategori.GELIR,
    ).order_by("ad")
    hata = None

    if request.method == "POST":
        kategori_id = request.POST.get("kategori")
        kategori = None
        if kategori_id:
            kategori = Kategori.objects.filter(
                id=kategori_id,
                kullanici=request.user,
                finans_turu=finans_turu,
                tur=Kategori.GELIR,
            ).first()

        if not kategori:
            hata = "Gelir kaydı için gelir kategorisi seçmelisin."
        else:
            gelir.tarih = request.POST["tarih"]
            gelir.aciklama = request.POST["aciklama"]
            gelir.tutar = request.POST["tutar"]
            gelir.finans_turu = finans_turu
            gelir.kategori = kategori.ad
            gelir.save()

            return redirect(f"/gelir-ekle/?finans_turu={finans_turu}")

    return render(request, "gelir_duzenle.html", {
        "gelir": gelir,
        "gelir_kategorileri": gelir_kategorileri,
        "hata": hata,
        **_finans_turu_context(finans_turu),
    })
